from flask import Flask, render_template, jsonify, request
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import smtplib

app = Flask(__name__)

def write_to_cell(ws, cell_coord, value):
    """Safely write to a cell, handling merged cells"""
    cell = ws[cell_coord]
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                top_left = merged_range.start_cell
                top_left.value = value
                return
    else:
        cell.value = value

def fill_excel_template(form_data):
    try:
        # Load template from local file
        template_path = "rirekisho_template.xlsx"
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Print form_data for debugging
        print("Received form data:", form_data)
        
        # Basic Information
        write_to_cell(ws, 'B4', form_data.get('furigana', ''))
        write_to_cell(ws, 'B5', form_data.get('name', ''))
        write_to_cell(ws, 'A6', form_data.get('birthdate', ''))
        write_to_cell(ws, 'C6', form_data.get('age', ''))
        write_to_cell(ws, 'A7', form_data.get('gender', ''))
        write_to_cell(ws, 'C7', form_data.get('family_structure', ''))
        write_to_cell(ws, 'A8', form_data.get('nationality', ''))
        write_to_cell(ws, 'C8', form_data.get('nearest_station', ''))
        write_to_cell(ws, 'B9', form_data.get('address', ''))
        write_to_cell(ws, 'B10', form_data.get('phone', ''))
        write_to_cell(ws, 'C10', form_data.get('email', ''))
        write_to_cell(ws, 'D10', form_data.get('social_media', ''))
        
        # Education History
        education = form_data.get('education', [])
        if isinstance(education, list):
            current_row = 13
            for edu in education:
                if isinstance(edu, dict):
                    write_to_cell(ws, f'A{current_row}', edu.get('entrance_date', ''))
                    write_to_cell(ws, f'B{current_row}', edu.get('graduation_date', ''))
                    write_to_cell(ws, f'C{current_row}', edu.get('school_name', ''))
                    write_to_cell(ws, f'D{current_row}', edu.get('department', ''))
                    current_row += 1
        
        # Work History
        work_history = form_data.get('work_history', [])
        if isinstance(work_history, list):
            current_row = 20
            for work in work_history:
                if isinstance(work, dict):
                    write_to_cell(ws, f'A{current_row}', work.get('start_date', ''))
                    write_to_cell(ws, f'B{current_row}', work.get('end_date', ''))
                    write_to_cell(ws, f'C{current_row}', work.get('company_name', ''))
                    write_to_cell(ws, f'D{current_row}', work.get('job_description', ''))
                    current_row += 1
        
        # Add creation date
        today = datetime.now()
        write_to_cell(ws, 'A50', f'作成日 {today.year}年{today.month}月{today.day}日')
        
        # Save as new file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"rirekisho_{timestamp}.xlsx"
        wb.save(output_filename)
        return output_filename
        
    except Exception as e:
        print(f"Error in fill_excel_template: {e}")
        print(f"Form data type: {type(form_data)}")
        print(f"Form data content: {form_data}")
        raise

def send_email(form_data):
    try:
        excel_filename = fill_excel_template(form_data)
        
        # Email configuration
        email = "baljir0901@gmail.com"
        password = "uvde dxqz useo xbdf"
        
        message = MIMEMultipart()
        message["From"] = email
        message["To"] = email
        message["Subject"] = "履歴書"
        
        # Attach Excel file
        with open(excel_filename, "rb") as f:
            excel_attachment = MIMEApplication(f.read(), _subtype="xlsx")
            excel_attachment.add_header(
                "Content-Disposition", "attachment", filename=excel_filename
            )
            message.attach(excel_attachment)
        
        # Send email
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(email, password)
            server.send_message(message)
        
        # Clean up
        os.remove(excel_filename)
        return True
        
    except Exception as e:
        print(f"Error in send_email: {e}")
        if 'excel_filename' in locals() and os.path.exists(excel_filename):
            os.remove(excel_filename)
        return False

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/submit-form', methods=['POST'])
def submit_form():
    try:
        form_data = request.get_json()
        print("Received form data:", form_data)  # Debug print
        
        if not isinstance(form_data, dict):
            print("Error: form_data is not a dictionary")
            return jsonify({"status": "error", "message": "Invalid form data format"})
        
        success = send_email(form_data)
        
        if success:
            return jsonify({"status": "success", "message": "履歴書が送信されました！"})
        else:
            return jsonify({"status": "error", "message": "送信に失敗しました"})
            
    except Exception as e:
        print(f"Error in submit_form: {e}")
        return jsonify({"status": "error", "message": str(e)})

if __name__ == '__main__':
    app.run(debug=True)